#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""SPM PivotTable -> du lieu dashboard (CHI nha may AH + QC An Ha).
Cau truc Pivot: ROW = Project>Zone>Workshop>WorkDate>QCWorkshop>QCDate>QCUser>memo
                COL = Stage (00.BOM,01.Fitup,02.Welding,03.Painting[,04.Final]) x {Qnty,Weight}
*** TU DONG DO CULL ***: doc 2 dong tieu de (Stage + Qnty/Weight) de tim DUNG cot Weight
   va so cap row-label, nen chay duoc CA layout cu (co 04.Final) LAN moi (khong Final / lech cot).
Xuong AH = ten Workshop CO CHUA 'AHx' (vd AH2.HHT1, TO5RHAH2, RHAH3...). Factory = AHx do.
San luong/BOM/du an/QC: suy TRUC TIEP tu dong tong phu Pivot -> khop 100%, khong bleed.
Waiting (cho QC accept) = SX cua to - phan da QC accept (QCWorkshop con trong = chua QC).
Dung: python spm_flatten.py [SPM.xlsx] [out_dir]
"""
import sys, os, re, json, csv, datetime, glob
from collections import Counter, defaultdict
try:
    import openpyxl
except ImportError:
    sys.exit("Thieu thu vien: chay  pip install openpyxl")
HERE=os.path.dirname(os.path.abspath(__file__))
def find_input():
    if len(sys.argv)>1 and os.path.exists(sys.argv[1]): return sys.argv[1]
    for pat in [os.path.join(HERE,'Update spm','*.xlsx'),os.path.join(HERE,'*.xlsx'),os.path.join(os.path.dirname(HERE),'Update spm','*.xlsx')]:
        for f in sorted(glob.glob(pat),key=os.path.getmtime,reverse=True):
            b=os.path.basename(f).lower()
            if 'spm' in b and not b.startswith('~'): return f
    sys.exit("Khong tim thay file SPM (*.xlsx) trong 'Update spm'.")
INP=find_input(); ODIR=(sys.argv[2] if len(sys.argv)>2 else HERE).strip().strip('"').strip() or HERE
CANON=['AH1','AH2','AH3','AH4','AH6','AH9']
PROC=['Fitup','Welding','Painting','Final']            # thu tu xuat ra (Final=0 neu file khong co)
STAGE_KW=[('BOM','bom'),('Fitup','fitup'),('Welding','welding'),('Painting','painting'),('Final','final')]
AHD=re.compile(r'AH[123469]')
def ah_parse(nm):
    U=str(nm).upper(); m=AHD.search(U)
    if not m: return None,None
    return m.group(0), (U[:m.start()]+U[m.end():]).strip('. -/') or '(chung)'
def clean(v):
    s=str(v).strip(); return s[:-6].strip() if s.endswith(' Total') else s
def cset(v):
    if v in (None,''): return None
    c=clean(v); return c if (c and c!='Total') else None
def shortpj(s):
    s=str(s).strip()
    s=re.sub(r'^((?:DGRP|[0-9][\w.\-/]*)\s+)+','',s,flags=re.I)
    s=re.sub(r'\s*[\(].*$','',s).strip()
    return (s[:26]).strip() or str(s)[:26]
def num(v): return v if isinstance(v,(int,float)) else 0
def sstr(v): return '' if v is None else str(v).strip()
try:
    wb=openpyxl.load_workbook(INP,read_only=True,data_only=True)
except Exception as e:
    sys.exit("\n[LOI] Khong mo duoc file SPM: %s\n   File '%s' co the BI HONG / CAT CUT hoac dang mo trong Excel.\n   -> DONG Excel, COPY/XUAT lai file day du, chep de vao 'Update spm' roi chay lai.\n   (Du lieu cu duoc GIU NGUYEN, khong ghi de.)"%(e, os.path.basename(INP)))
ws=wb[wb.sheetnames[0]]
allrows=[list(r) for r in ws.iter_rows(values_only=True)]
wb.close()
maxc=max((len(r) for r in allrows[:60]),default=0)
allrows=[r+[None]*(max(maxc,22)-len(r)) for r in allrows]

# ===== TU DONG DO CULL: tim dong tieu de Stage + dong Qnty/Weight =====
HR=None
for i,r in enumerate(allrows[:40]):
    if any('bom' in sstr(c).lower() for c in r) and any('fitup' in sstr(c).lower() for c in r):
        HR=i; break
if HR is None:
    sys.exit("\n[LOI] Khong tim thay dong tieu de cong doan (00.BOM / 01.Fitup...).\n   -> File SPM co the sai dinh dang. Hay xuat lai PivotTable dung mau.")
SUB=HR+1
hdr=allrows[HR]; sub=allrows[SUB]
STG_W={}; STG_HDR={}              # stage -> cot Weight / cot tieu de (= cot Qnty dau cong doan)
for key,kw in STAGE_KW:
    for j,c in enumerate(hdr):
        if kw in sstr(c).lower():
            STG_HDR[key]=j
            wc=None
            for cc in (j,j+1,j+2):
                if cc<len(sub) and sstr(sub[cc]).lower()=='weight': wc=cc; break
            if wc is None: wc=j+1     # mac dinh Weight = cot ke phai
            STG_W[key]=wc; break
if 'BOM' not in STG_W or 'Fitup' not in STG_W:
    sys.exit("\n[LOI] Khong do duoc cot Weight cua BOM/Fitup. File sai mau.")
BOMW=STG_W['BOM']
ND=min(STG_HDR.values())          # so cot row-label = cot tieu de cong doan dau (BOM) = noi vung du lieu bat dau

if ND<4:
    sys.exit("""
[LOI] File SPM xuat SAI KIEU PIVOT: chi co %d cot nhan dong (Project, Zone).
   Cac truong Workshop / WorkDate / QCWorkshop / QCDate / QCUser dang nam o vung
   FILTER hoac bi bo khoi Pivot -> khong co du lieu chi tiet de len dashboard.
   -> Mo lai Pivot trong SPM, keo du cac truong vao vung ROWS theo thu tu:
      Project > Zone > Workshop > WorkDate > QCWorkshop > QCDate > QCUser (> memo)
      COLUMNS = Stage, VALUES = Qnty + Weight. Xuat lai file roi chay lai.
   (Du lieu cu tren dashboard duoc GIU NGUYEN, khong ghi de.)"""%ND)
DATA0=SUB+1                       # chi so dong du lieu dau tien (0-based)
PRESENT=[s for s in PROC if s in STG_W]
def gw(row,stage):                # weight cua 1 cong doan (0 neu file khong co)
    c=STG_W.get(stage)
    return num(row[c]) if (c is not None and c<len(row)) else 0
def gq(row,stage):                # so luong (Qnty) cua 1 cong doan
    c=STG_HDR.get(stage)
    return int(round(num(row[c]))) if (c is not None and c<len(row)) else 0
# ===== TU DONG NHAN DIEN VI TRI CAP row-label theo NOI DUNG =====
# (Pivot co the doi thu tu WorkDate/QCWorkshop/QCUser/QCDate giua cac lan xuat)
import collections as _co
_samp=_co.defaultdict(list)
for _row in allrows[DATA0:DATA0+8000]:
    for _c in range(1,ND):
        _v=_row[_c] if _c<len(_row) else None
        if _v not in (None,''):
            _sv=str(_v).strip()
            if _sv.endswith('Total'): _sv=_sv[:-6].strip()
            if _sv and _sv!='Total': _samp[_c].append(_sv)
def _clsfy(vals):
    if not vals: return 'empty'
    n=len(vals)
    nd=sum(1 for v in vals if re.match(r'^\d{1,2}/\d{1,2}/\d{4}$',v) or re.match(r'^\d{4}-\d{2}-\d{2}',v))
    qc=sum(1 for v in vals if re.search(r'\.(QC|VP)\b',v,re.I))
    us=sum(1 for v in vals if re.match(r'^[a-z][a-z0-9._]{1,14}$',v))
    sg=sum(1 for v in vals if re.match(r'^\d\d?\.(Waiting|Accept|Reject)',v,re.I))
    if sg/n>0.5: return 'status'
    if nd/n>0.6: return 'date'
    if qc/n>0.5: return 'qcws'
    if us/n>0.5: return 'user'
    return 'other'
_clsm={_c:_clsfy(_samp[_c]) for _c in range(1,ND)}
_dateC=sorted([c for c in _clsm if _clsm[c]=='date'])
_qcwsC=sorted([c for c in _clsm if _clsm[c]=='qcws'])
_userC=sorted([c for c in _clsm if _clsm[c]=='user'])
L_PROJ,L_ZONE,L_WS=0,1,2
L_WD=_dateC[0] if _dateC else 3
L_QCDATE=(_dateC[-1] if len(_dateC)>=2 else None)
L_QCWS=_qcwsC[0] if _qcwsC else 4
L_QCU=_userC[0] if _userC else None   # khong co cot nguoi kiem -> de trong, KHONG lay nham memo
_statC=sorted([c for c in _clsm if _clsm[c]=='status']); L_STATUS=_statC[0] if _statC else None
L_MEMO=ND-1
print("   [DO CAP] WorkDate=c%s QCWorkshop=c%s QCUser=c%s QCDate=c%s Status=c%s"%(L_WD,L_QCWS,L_QCU,L_QCDATE,L_STATUS))

ctxp=None; cur_xu=cur_to=None
projBOM={}; byTo={}; byProj={}; byToProj=defaultdict(lambda:defaultdict(float)); qcW=defaultdict(lambda:defaultdict(float))
waitTo=defaultdict(float); qcAllByTo=defaultdict(float); flatAgg=defaultdict(lambda:[0.0,0.0,0.0,0.0])
fctx=[None]*ND; flatFull=[]; bomAH=defaultdict(float)
for row in allrows[DATA0:]:
    if ND>len(row): row=row+[None]*(ND-len(row))
    # ===== FLAT row-level: chi AH (Workshop chua AHx), kem QC user + phan loai QC + BOM =====
    fsub=False
    for _i in range(ND):
        _cv=row[_i]
        if _cv not in (None,''):
            _sv=str(_cv).strip()
            if _sv.endswith('Total'): fsub=True; fctx[_i]=_sv[:-6].strip()
            else: fctx[_i]=_sv
            for _j in range(_i+1,ND): fctx[_j]=None
    if not fsub:
        _bom=num(row[BOMW]); _fw=[gw(row,s) for s in PROC]   # F,W,P,Final theo thu tu PROC
        if _bom or any(_fw):
            _fx,_ft=ah_parse(fctx[L_WS] or ''); _qc=(fctx[L_QCWS] or '')
            if _fx in CANON:
                _statusRaw=(fctx[L_STATUS] if L_STATUS is not None else '') or '';_stv=_statusRaw.lower()
                if _stv:
                    _pl='Rớt' if 'reject' in _stv else ('Đã QC (AH)' if 'accept' in _stv else 'Chưa QC')
                else:
                    _pl='Đã QC (AH)' if _qc.upper().startswith('AH') else ('Chưa QC' if not _qc.strip() else 'QC khác')
                _memo=(fctx[L_MEMO] if (L_MEMO is not None and L_MEMO!=L_STATUS and L_MEMO not in (L_QCU,L_QCDATE,L_QCWS,L_WD)) else '') or ''
                flatFull.append([fctx[L_PROJ] or '',fctx[L_ZONE] or '',_fx,_ft,fctx[L_QCDATE] or '',_qc,((fctx[L_QCU] if L_QCU is not None else '') or ''),_pl,_statusRaw]+[round(x/1000,3) for x in _fw]+[fctx[L_WD] or '']+[gq(row,'Fitup'),gq(row,'Welding'),gq(row,'Painting'),_memo])
                bomAH[fctx[L_PROJ] or '']+=_bom
    c0=cset(row[L_PROJ])
    if c0 is not None: ctxp=c0
    if c0 is not None and num(row[BOMW])>0: projBOM[ctxp]=max(projBOM.get(ctxp,0),num(row[BOMW]))
    rawws=row[L_WS]
    if rawws not in (None,'') and str(rawws).strip().endswith('Total'):
        nm=clean(rawws); xu,to=ah_parse(nm)
        if xu and xu in CANON:
            stg={s:gw(row,s) for s in PROC}
            a=byTo.setdefault((xu,to),{s:0.0 for s in PROC}); p=byProj.setdefault(ctxp,{s:0.0 for s in PROC})
            for s in PROC: a[s]+=stg[s]; p[s]+=stg[s]
            tw=sum(stg.values())
            if tw>0:
                byToProj[(xu,to)][ctxp]+=tw
                fa=flatAgg[(ctxp,xu,to)]
                for _i,_s in enumerate(PROC): fa[_i]+=stg[_s]
    vws=cset(row[L_WS])
    if vws is not None:
        _f,_t=ah_parse(vws)
        if _f in CANON: cur_xu,cur_to=_f,_t
        else: cur_xu=cur_to=None
    rawq=row[L_QCWS]
    if rawq not in (None,'') and str(rawq).strip().endswith('Total') and cur_to is not None:
        q=clean(rawq); _w=sum(gw(row,s) for s in PROC)
        qcAllByTo[(cur_xu,cur_to)]+=_w
        if q and q.upper().startswith('AH'): qcW[(cur_xu,cur_to)][q]+=_w
# QC USER theo to (tu flatFull): [2]Xuong [3]To [6]QCUser [9..12] F/W/P/Final
qcUserByTo=defaultdict(lambda:defaultdict(float))
for _fr in flatFull:
    _u=str(_fr[6]).strip()
    if _u: qcUserByTo[(_fr[2],_fr[3])][_u]+=(_fr[9]+_fr[10]+_fr[11]+_fr[12])
def monthkey(wd):
    wd=str(wd).strip()
    m=re.match(r'(\d{4})-(\d{2})',wd)
    if m: return m.group(1)+'-'+m.group(2)
    m=re.match(r'(\d{1,2})/(\d{1,2})/(\d{4})',wd)
    if m: return '%s-%02d'%(m.group(3),int(m.group(1)))
    return None
byMonth=defaultdict(lambda:[0.0,0.0,0.0])
for _fr in flatFull:
    _mk=monthkey(_fr[4])
    if not _mk: continue
    byMonth[_mk][0]+=_fr[9]; byMonth[_mk][1]+=_fr[10]; byMonth[_mk][2]+=_fr[11]
byMonthL=[{'m':k,'f':round(v[0],1),'w':round(v[1],1),'p':round(v[2],1)} for k,v in sorted(byMonth.items())]
qcStats={}
for _fr in flatFull:
    if _fr[7]!='Đã QC (AH)': continue
    _u=str(_fr[6]).strip()
    if not _u: continue
    _F,_W,_P,_Fin=_fr[9],_fr[10],_fr[11],_fr[12]; _tot=_F+_W+_P+_Fin
    _sx=qcStats.setdefault(_u,{'kl':0.0,'n':0,'f':0.0,'w':0.0,'p':0.0,'tos':defaultdict(float)})
    _sx['kl']+=_tot; _sx['n']+=1; _sx['f']+=_F; _sx['w']+=_W; _sx['p']+=_P; _sx['tos'][_fr[2]+'.'+_fr[3]]+=_tot
byQC=[]
for _u,_sx in qcStats.items():
    _tos=sorted(_sx['tos'].items(),key=lambda kv:-kv[1])[:4]
    byQC.append({'u':_u,'kl':round(_sx['kl'],1),'n':_sx['n'],'f':round(_sx['f'],1),'w':round(_sx['w'],1),'p':round(_sx['p'],1),'tos':[t for t,_ in _tos]})
byQC.sort(key=lambda x:-x['kl'])
for _k,_sd in byTo.items():
    waitTo[_k]=max(0.0, sum(_sd.values()) - qcAllByTo.get(_k,0))
t3=lambda x: round(x/1000,1)
def status(pct,bom):
    if bom==0: return 'Không BOM'
    if pct>=95: return 'Hoàn thành'
    if pct>=50: return 'Đang hoàn thiện'
    if pct>0: return 'Còn tồn đọng'
    return 'Chưa sơn'
byProjL=[]
for pr,p in byProj.items():
    if not pr: continue
    bom=projBOM.get(pr,0); pct=round(p['Painting']/bom*100) if bom else 0
    byProjL.append({'proj':pr,'bom':t3(bom),'st':{s:t3(p[s]) for s in PROC},'ton':{s:t3(max(0,bom-p[s])) for s in PROC},'pct':pct,'status':status(pct,bom),'wait':t3(0)})
byProjL.sort(key=lambda x:-x['ton']['Painting'])
byToOut=[]
for (xu,to),v in byTo.items():
    k=(xu,to)
    pj=sorted(byToProj.get(k,{}).items(),key=lambda kv:-kv[1])[:3]; qc=sorted(qcUserByTo.get(k,{}).items(),key=lambda kv:-kv[1])[:3]
    byToOut.append({'x':xu,'to':to,'f':t3(v['Fitup']),'w':t3(v['Welding']),'p':t3(v['Painting']),'wait':t3(waitTo.get(k,0)),'proj':[[shortpj(pp),t3(ww)] for pp,ww in pj if ww>0],'qc':[q for q,_ in qc] or ['—']})
byToOut.sort(key=lambda d:(d['x'],-(d['f']+d['w']+d['p'])))
byXacc={}
for (xu,to),v in byTo.items():
    a=byXacc.setdefault(xu,{'f':0.0,'w':0.0,'p':0.0,'wait':0.0})
    a['f']+=v['Fitup']; a['w']+=v['Welding']; a['p']+=v['Painting']; a['wait']+=waitTo.get((xu,to),0)
byXL=[{'x':xu,'f':t3(a['f']),'w':t3(a['w']),'p':t3(a['p']),'wait':t3(a['wait'])} for xu,a in sorted(byXacc.items())]
statusCounts=dict(Counter(r['status'] for r in byProjL))
topTon=[[shortpj(r['proj']),r['ton']['Painting']] for r in byProjL if r['bom']>0][:6]
TB=sum(projBOM.get(p,0) for p in byProj if p)
totals={'bom':t3(TB),'fitup':t3(sum(v['Fitup'] for v in byTo.values())),'welding':t3(sum(v['Welding'] for v in byTo.values())),'painting':t3(sum(v['Painting'] for v in byTo.values())),'wait':t3(sum(waitTo.values()))}
out={'updated':datetime.date.today().strftime('%d/%m/%Y'),'source':os.path.basename(INP),'hasFinal':('Final' in STG_W),'hasWaiting':True,'waitMethod':'SX - da QC accept','totals':totals,'byX':byXL,'status':statusCounts,'topTon':topTon,'byMonth':byMonthL,'byQC':byQC,'byTo':byToOut,'byProject':byProjL}
_sx=totals['fitup']+totals['welding']+totals['painting']
if _sx<=0 or len(flatFull)==0:
    sys.exit("\n[LOI] Doc duoc cau truc nhung TONG SAN LUONG = 0 tan (file co the mat gia tri cache / sai dinh dang).\n   -> Mo file SPM bang Excel, nhan Ctrl+S de luu lai (tao cache gia tri), roi chay lai.\n   (Du lieu cu duoc GIU NGUYEN, khong ghi de.)")
# ===== CAP NHAT = THAY SACH: luon ghi de hoan toan theo file moi (KHONG giu du lieu cu) =====
_lcf=os.path.join(ODIR,'DuLieu_SPM','.last_count')
try: _lastn=int(open(_lcf,encoding='utf-8').read().strip())
except Exception: _lastn=0
# (ban cu luon duoc tu dong sao luu trong DuLieu_SPM/ truoc khi ghi de -> khong mat vinh vien)
_wds=[fr[13] for fr in flatFull if fr[13]]
def _mmd(v):
    m=re.match(r'(\d{1,2})/(\d{1,2})/(\d{4})',str(v)); return (int(m.group(3)),int(m.group(1)),int(m.group(2))) if m else (0,0,0)
if _wds:
    print('   [THAY SACH] Nap MOI %d dong | WorkDate %s -> %s (ghi de toan bo, KHONG giu ban cu)'%(len(flatFull),min(_wds,key=_mmd),max(_wds,key=_mmd)))
if _lastn>=200 and len(flatFull)<_lastn*0.6:
    print('   [i] Luu y: lan nay it hon lan truoc (%d vs %d dong). Neu CO Y export loc thi bo qua; neu khong, kiem tra lai bo loc roi chay lai.'%(len(flatFull),_lastn))
open(os.path.join(ODIR,'spm_data.js'),'w',encoding='utf-8').write('window.SPM_DATA='+json.dumps(out,ensure_ascii=False)+';')
with open(os.path.join(ODIR,'spm_project.csv'),'w',newline='',encoding='utf-8-sig') as fp:
    w=csv.writer(fp); w.writerow(['Project','BOM_caduan_t','Fitup_t','Welding_t','Painting_t','Final_t','Ton_Fitup','Ton_Welding','Ton_Painting','Ton_Final','PhanTram','TrangThai'])
    for r in byProjL: w.writerow([r['proj'],r['bom']]+[r['st'][s] for s in PROC]+[r['ton'][s] for s in PROC]+[r['pct'],r['status']])
with open(os.path.join(ODIR,'spm_flat.csv'),'w',newline='',encoding='utf-8-sig') as fp:
    w=csv.writer(fp); w.writerow(['Project','Zone','Xuong','To','QCDate','QCWorkshop','QCUser','PhanLoaiQC','Status','Fitup_t','Welding_t','Painting_t','Final_t','WorkDate','Qnty_F','Qnty_W','Qnty_P','Memo'])
    w.writerows(flatFull)
with open(os.path.join(ODIR,'spm_to.csv'),'w',newline='',encoding='utf-8-sig') as fp:
    w=csv.writer(fp); w.writerow(['Project','Xuong','To','Fitup_t','Welding_t','Painting_t','Final_t','Tong_t'])
    for (pr,xu,to),stv in sorted(flatAgg.items(),key=lambda kv:(kv[0][0],-sum(kv[1]))):
        w.writerow([pr,xu,to]+[round(v/1000,3) for v in stv]+[round(sum(stv)/1000,3)])
# ----- qcdata.js cho dashboard_spm_qc.html (TU NAP - khoi nhung tay) -----
def _isoq(x):
    m=re.match(r'(\d{1,2})/(\d{1,2})/(\d{4})',str(x or ''))
    return '%s-%02d-%02d'%(m.group(3),int(m.group(1)),int(m.group(2))) if m else ''
_P=[];_X=[];_T=[];_Q=[];_PL=[];_Pm={};_Xm={};_Tm={};_Qm={};_PLm={}
def _ixq(mp,lst,v):
    if v not in mp: mp[v]=len(lst);lst.append(v)
    return mp[v]
_R=[]
for fr in flatFull:
    _R.append([_ixq(_Pm,_P,fr[0]),_ixq(_Xm,_X,fr[2]),_ixq(_Tm,_T,fr[3]),_isoq(fr[4]),_isoq(fr[13]),
               _ixq(_Qm,_Q,(fr[6] or '—')),_ixq(_PLm,_PL,fr[7]),
               round(fr[9],3),round(fr[10],3),round(fr[11],3),int(fr[14]),int(fr[15]),int(fr[16]),fr[17]])
_projq=[{'name':r['proj'],'bom':r['bom'],'f':r['st']['Fitup'],'w':r['st']['Welding'],'p':r['st']['Painting'],
         'tf':r['ton']['Fitup'],'tw':r['ton']['Welding'],'tp':r['ton']['Painting'],'pct':r['pct'],'status':r['status']} for r in byProjL]
_qcobj={'updated':out['updated'],'source':os.path.basename(INP),'P':_P,'X':_X,'T':_T,'Q':_Q,'PL':_PL,'rows':_R,'proj':_projq,
        'check':{'f':round(sum(x[7] for x in _R),1),'w':round(sum(x[8] for x in _R),1),'p':round(sum(x[9] for x in _R),1)}}
_qc_txt='window.QCDATA='+json.dumps(_qcobj,ensure_ascii=False,separators=(',',':'))+';'
_qc_tmp=os.path.join(ODIR,'qcdata.js.tmp')
with open(_qc_tmp,'w',encoding='utf-8') as _qf:
    _qf.write(_qc_txt); _qf.flush(); os.fsync(_qf.fileno())
# doc lai file tam, chi thay the khi nguyen ven (chong ghi cut / OneDrive can thiep)
_qc_back=open(_qc_tmp,encoding='utf-8').read()
if len(_qc_back)==len(_qc_txt) and _qc_back.rstrip().endswith(';'):
    os.replace(_qc_tmp,os.path.join(ODIR,'qcdata.js'))
else:
    print('    [!] GHI LOI: file tam bi cut (%d/%d) — GIU NGUYEN qcdata.js cu, chay lai!'%(len(_qc_back),len(_qc_txt)))
_qcp=os.path.join(ODIR,'qcdata.js')
try:
    _chk=open(_qcp,encoding='utf-8').read(); _co=json.loads(_chk[_chk.index('{'):_chk.rindex('}')+1]); _cn=len(_co.get('rows',[]))
    if _chk.rstrip().endswith(';') and _cn==len(_R): print("    KIEM TRA: qcdata.js HOP LE - %d dong (KHONG bi cut)."%_cn)
    else: print("    [!] CANH BAO: qcdata.js nghi LOI (dong %d/%d) - dung dung, chay lai!"%(_cn,len(_R)))
except Exception as _qe:
    print("    [!] CANH BAO: qcdata.js KHONG doc lai duoc -",_qe)
try:
    os.makedirs(os.path.join(ODIR,'DuLieu_SPM'),exist_ok=True)
    open(os.path.join(ODIR,'DuLieu_SPM','.last_count'),'w',encoding='utf-8').write(str(len(flatFull)))
except Exception: pass
import shutil as _sh
_arch=os.path.join(ODIR,'DuLieu_SPM'); os.makedirs(_arch,exist_ok=True)
_ts=datetime.datetime.now().strftime('%Y%m%d_%H%M')
# Chi tao ban backup theo ngay-gio khi SPM THUC SU doi (tranh rac khi chay lich lap moi 30')
try:
    _ist=os.stat(INP); _sig='%d_%d_%d'%(int(_ist.st_mtime),_ist.st_size,len(flatFull))
except Exception:
    _sig=str(len(flatFull))
_sigf=os.path.join(_arch,'.last_sig'); _changed=True
try:
    if os.path.exists(_sigf) and open(_sigf,encoding='utf-8').read().strip()==_sig: _changed=False
except Exception: pass
for _fn in ('spm_data.js','spm_project.csv','spm_flat.csv','spm_to.csv'):
    _src=os.path.join(ODIR,_fn)
    if os.path.exists(_src):
        _sh.copy2(_src,os.path.join(_arch,_fn))
        if _changed:
            _st,_ex=os.path.splitext(_fn); _sh.copy2(_src,os.path.join(_arch,_st+'_'+_ts+_ex))
try: open(_sigf,'w',encoding='utf-8').write(_sig)
except Exception: pass
print("OK  do cot tu dong: BOM_w=%d  Fitup_w=%d  ...  Final=%s  | row-label=%d cap  | data tu dong %d"%(BOMW,STG_W['Fitup'],('co' if 'Final' in STG_W else 'KHONG'),ND,DATA0+1))
print("    spm_flat.csv = %d dong (TOAN BO AH, co cot PhanLoaiQC + QC user) | %s"%(len(flatFull),("Da luu ban moi DuLieu_SPM/ (%s)"%_ts) if _changed else "SPM khong doi -> khong tao backup moi"))
